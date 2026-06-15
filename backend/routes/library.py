"""Bibliothèque centrale de documents (factures, devis, etc.)

Stockage : Emergent Object Storage (persistant entre les déploiements).
"""
import logging
import re
from pathlib import Path
from typing import Optional, Literal, List
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Query, Body
from fastapi.responses import Response
from pydantic import BaseModel

import deps
from storage import APP_PREFIX, put_object, get_object

logger = logging.getLogger("crm.library")
router = APIRouter()

DocTypeLiteral = Literal["devis_signe", "attestation", "facture", "justificatif_paiement", "autre"]

DOC_TYPE_LABEL = {
    "devis_signe": "Devis signé",
    "attestation": "Attestation",
    "facture": "Facture",
    "justificatif_paiement": "Justificatif de paiement",
    "autre": "Autre",
}


class AttachPayload(BaseModel):
    dossier_id: str


# ---------------------------------------------------------------------------
# Auto-rattachement PDF facture → apprenant
# ---------------------------------------------------------------------------

# Reconnaît BA-1077, B-A1077, BA1077, B_A1077, ba 1077.
# Pattern strict : 1-3 lettres + (séparateur optionnel + 1 lettre optionnelle) + séparateur optionnel + 3-8 chiffres + FIN.
_INVOICE_NUM_RE = re.compile(r"^\s*([A-Za-z]{1,3})[\s\-_/.]?([A-Za-z]?)[\s\-_/.]?(\d{3,8})\s*$")


def _normalize_invoice_number(value: Optional[str]) -> Optional[str]:
    """Normalise un n° de facture (ou nom de fichier) en clé canonique.

    Exemples :
        "BA-1077.pdf"  → "BA1077"
        "B-A1077.pdf"  → "BA1077"
        "BA 1077"      → "BA1077"
        "ba_1077"      → "BA1077"
        "Rapport.xlsx" → None (trop de lettres avant les chiffres)
        "Facture-BA-1077.pdf" → None (préfixe « Facture » non standard)
    """
    if not value:
        return None
    stem = Path(value).stem.strip()  # retire l'extension si fichier
    m = _INVOICE_NUM_RE.match(stem)
    if not m:
        return None
    cleaned = (m.group(1) + m.group(2) + m.group(3)).upper()
    return cleaned


async def _match_apprenant_for_filename(filename: Optional[str]) -> dict:
    """Tente d'identifier l'apprenant cible d'un PDF à partir de son nom de fichier.

    Retourne un dict avec :
      - status: 'ok' | 'unparseable' | 'invoice_not_found' | 'no_apprenant'
      - apprenant_id (si ok)
      - apprenant_nom, apprenant_prenom (si ok)
      - numero_facture, numero_dossier (si trouvés)
      - reason (message lisible)
    """
    key = _normalize_invoice_number(filename)
    if not key:
        return {"status": "unparseable", "reason": "Nom de fichier non reconnu comme n° de facture"}

    # Recherche TOUS les n° facture qui, une fois normalisés, matchent la clé
    facture = None
    async for f in deps.db.factures_cpf.find(
        {"numero_facture": {"$nin": [None, ""]}},
        {"_id": 0, "id": 1, "numero_facture": 1, "numero_dossier": 1},
    ):
        if _normalize_invoice_number(f.get("numero_facture")) == key:
            facture = f
            break

    if not facture:
        return {"status": "invoice_not_found", "reason": f"Aucune facture CPF ne correspond au n° « {key} »", "numero_facture": key}

    numero_dossier = facture.get("numero_dossier")
    if not numero_dossier:
        return {
            "status": "no_apprenant",
            "reason": f"Facture {facture.get('numero_facture')} sans n° de dossier CPF",
            "numero_facture": facture.get("numero_facture"),
        }

    apprenant = await deps.db.apprenants.find_one(
        {"dossier_cpf": numero_dossier},
        {"_id": 0, "id": 1, "nom": 1, "prenom": 1},
    )
    if not apprenant:
        return {
            "status": "no_apprenant",
            "reason": f"Aucun apprenant avec dossier CPF « {numero_dossier} »",
            "numero_facture": facture.get("numero_facture"),
            "numero_dossier": numero_dossier,
        }

    return {
        "status": "ok",
        "apprenant_id": apprenant["id"],
        "apprenant_nom": apprenant.get("nom"),
        "apprenant_prenom": apprenant.get("prenom"),
        "numero_facture": facture.get("numero_facture"),
        "numero_dossier": numero_dossier,
    }


async def _fetch_bytes(doc: dict) -> tuple:
    """Récupère les bytes + content_type d'un document depuis le stockage objet."""
    path = doc.get("storage_path")
    if not path:
        raise HTTPException(404, "Fichier introuvable (storage_path manquant — document antérieur à la migration objet)")
    try:
        data, ctype = await get_object(path)
    except Exception:
        logger.exception("Échec de lecture du stockage objet")
        raise HTTPException(502, "Stockage indisponible, réessayez dans un instant")
    return data, ctype or doc.get("content_type", "application/octet-stream")


async def _enrich_with_stagiaire(docs: List[dict]) -> List[dict]:
    dossier_ids = list({d["dossier_id"] for d in docs if d.get("dossier_id")})
    apprenant_ids = list({d["apprenant_id"] for d in docs if d.get("apprenant_id")})
    mapping_dossier = {}
    mapping_apprenant = {}
    if dossier_ids:
        async for d in deps.db.dossiers.find(
            {"id": {"$in": dossier_ids}}, {"_id": 0, "id": 1, "nom": 1, "prenom": 1, "financeur_type": 1}
        ):
            mapping_dossier[d["id"]] = d
    if apprenant_ids:
        async for a in deps.db.apprenants.find(
            {"id": {"$in": apprenant_ids}}, {"_id": 0, "id": 1, "nom": 1, "prenom": 1, "dossier_cpf": 1}
        ):
            mapping_apprenant[a["id"]] = a
    for doc in docs:
        did = doc.get("dossier_id")
        if did and did in mapping_dossier:
            d = mapping_dossier[did]
            doc["stagiaire"] = {"id": d["id"], "nom": d.get("nom"), "prenom": d.get("prenom"), "financeur_type": d.get("financeur_type")}
        else:
            doc["stagiaire"] = None
        aid = doc.get("apprenant_id")
        if aid and aid in mapping_apprenant:
            a = mapping_apprenant[aid]
            doc["apprenant"] = {"id": a["id"], "nom": a.get("nom"), "prenom": a.get("prenom"), "dossier_cpf": a.get("dossier_cpf")}
        else:
            doc["apprenant"] = None
    return docs


@router.get("/library")
async def list_library(
    scope: str = Query("all", pattern="^(all|unattached|attached)$"),
    type: Optional[str] = None,
    q: Optional[str] = None,
    user: dict = Depends(deps.get_current_user),
):
    import re as _re
    base_query: dict = {}
    if scope == "unattached":
        base_query["$and"] = [
            {"$or": [{"dossier_id": None}, {"dossier_id": {"$exists": False}}]},
            {"$or": [{"apprenant_id": None}, {"apprenant_id": {"$exists": False}}]},
        ]
    elif scope == "attached":
        base_query["$or"] = [
            {"dossier_id": {"$nin": [None, ""]}},
            {"apprenant_id": {"$nin": [None, ""]}},
        ]
    if type and type in DOC_TYPE_LABEL:
        base_query["type"] = type
    docs = await deps.db.dossier_documents.find(base_query, {"_id": 0}).sort("uploaded_at", -1).to_list(5000)
    if q:
        term = q.strip().lower()
        dossier_match_ids = set()
        async for d in deps.db.dossiers.find(
            {"$or": [
                {"nom": {"$regex": _re.escape(q), "$options": "i"}},
                {"prenom": {"$regex": _re.escape(q), "$options": "i"}},
                {"financeur_nom": {"$regex": _re.escape(q), "$options": "i"}},
            ]},
            {"_id": 0, "id": 1},
        ):
            dossier_match_ids.add(d["id"])
        cpf_dossier_numbers = set()
        async for f in deps.db.factures_cpf.find(
            {"$or": [
                {"numero_facture": {"$regex": _re.escape(q), "$options": "i"}},
                {"numero_dossier": {"$regex": _re.escape(q), "$options": "i"}},
            ]},
            {"_id": 0, "numero_dossier": 1},
        ):
            if f.get("numero_dossier"):
                cpf_dossier_numbers.add(f["numero_dossier"])
        if cpf_dossier_numbers:
            async for d in deps.db.dossiers.find(
                {"financeur_nom": {"$in": list(cpf_dossier_numbers)}},
                {"_id": 0, "id": 1},
            ):
                dossier_match_ids.add(d["id"])
        filtered = []
        for doc in docs:
            fname = (doc.get("original_filename") or "").lower()
            did = doc.get("dossier_id")
            if term in fname or (did and did in dossier_match_ids):
                filtered.append(doc)
        docs = filtered
    return await _enrich_with_stagiaire(docs)


@router.get("/library-stats")
async def library_stats(user: dict = Depends(deps.get_current_user)):
    total = await deps.db.dossier_documents.count_documents({})
    attached = await deps.db.dossier_documents.count_documents({
        "$or": [
            {"dossier_id": {"$nin": [None, ""]}},
            {"apprenant_id": {"$nin": [None, ""]}},
        ]
    })
    by_type = {}
    for t in DOC_TYPE_LABEL.keys():
        by_type[t] = await deps.db.dossier_documents.count_documents({"type": t})
    total_size = 0
    async for d in deps.db.dossier_documents.find({}, {"_id": 0, "size": 1}):
        total_size += d.get("size") or 0
    return {"total": total, "attached": attached, "unattached": total - attached, "by_type": by_type, "total_size_bytes": total_size}


@router.post("/library/upload")
async def upload_library(
    file: UploadFile = File(...),
    type: DocTypeLiteral = Form("facture"),
    dossier_id: Optional[str] = Form(None),
    user: dict = Depends(deps.get_current_user),
):
    content = await file.read()
    if not content:
        raise HTTPException(400, "Fichier vide")
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(400, "Fichier trop volumineux (max 25 Mo)")
    if dossier_id:
        if not await deps.db.dossiers.find_one({"id": dossier_id}, {"_id": 0, "id": 1}):
            raise HTTPException(400, "Dossier stagiaire introuvable")
    doc_id = deps.new_id()
    ext = Path(file.filename or "").suffix.lstrip(".") or "bin"
    path = f"{APP_PREFIX}/library/{doc_id}.{ext}"
    try:
        result = await put_object(path, content, file.content_type or "application/octet-stream")
    except Exception:
        logger.exception("Échec d'upload vers le stockage objet")
        raise HTTPException(502, "Stockage indisponible, réessayez dans un instant")
    document = {
        "id": doc_id,
        "dossier_id": dossier_id,
        "apprenant_id": None,
        "type": type,
        "filename": path.split("/")[-1],  # back-compat affichage
        "original_filename": file.filename or path.split("/")[-1],
        "content_type": file.content_type or "application/octet-stream",
        "size": result.get("size", len(content)),
        "storage_path": result.get("path", path),
        "uploaded_at": deps.now_utc().isoformat(),
    }
    # Auto-rattachement à l'apprenant si c'est une facture (basé sur n° dans le nom de fichier)
    # NB : on n'auto-rattache que les PDF/images, jamais les Excel/Word/CSV (mêmes règles que /library/auto-attach)
    if type == "facture" and not dossier_id:
        skip_exts = {"xlsx", "xls", "xlsm", "csv", "doc", "docx"}
        upload_ext = Path(file.filename or "").suffix.lower().lstrip(".")
        if upload_ext not in skip_exts:
            match = await _match_apprenant_for_filename(file.filename)
            if match.get("status") == "ok":
                document["apprenant_id"] = match["apprenant_id"]
                document["auto_attached"] = True
                document["auto_attach_meta"] = {
                    "numero_facture": match.get("numero_facture"),
                    "numero_dossier": match.get("numero_dossier"),
                }
    await deps.db.dossier_documents.insert_one(dict(document))
    document.pop("_id", None)
    enriched = await _enrich_with_stagiaire([document])
    return enriched[0]


@router.patch("/library/{document_id}/attach")
async def attach_document(document_id: str, payload: AttachPayload, user: dict = Depends(deps.get_current_user)):
    if not await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0, "id": 1}):
        raise HTTPException(404, "Document introuvable")
    if not await deps.db.dossiers.find_one({"id": payload.dossier_id}, {"_id": 0, "id": 1}):
        raise HTTPException(400, "Dossier stagiaire introuvable")
    await deps.db.dossier_documents.update_one({"id": document_id}, {"$set": {"dossier_id": payload.dossier_id}})
    updated = await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0})
    enriched = await _enrich_with_stagiaire([updated])
    return enriched[0]


class AttachApprenantPayload(BaseModel):
    apprenant_id: str


@router.patch("/library/{document_id}/attach-apprenant")
async def attach_document_apprenant(document_id: str, payload: AttachApprenantPayload, user: dict = Depends(deps.get_current_user)):
    """Rattachement manuel d'un document de la bibliothèque à un apprenant."""
    if not await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0, "id": 1}):
        raise HTTPException(404, "Document introuvable")
    if not await deps.db.apprenants.find_one({"id": payload.apprenant_id}, {"_id": 0, "id": 1}):
        raise HTTPException(400, "Apprenant introuvable")
    await deps.db.dossier_documents.update_one(
        {"id": document_id}, {"$set": {"apprenant_id": payload.apprenant_id}}
    )
    updated = await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0})
    enriched = await _enrich_with_stagiaire([updated])
    return enriched[0]


@router.patch("/library/{document_id}/detach-apprenant")
async def detach_document_apprenant(document_id: str, user: dict = Depends(deps.get_current_user)):
    if not await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0, "id": 1}):
        raise HTTPException(404, "Document introuvable")
    await deps.db.dossier_documents.update_one(
        {"id": document_id},
        {"$set": {"apprenant_id": None, "auto_attached": False}, "$unset": {"auto_attach_meta": ""}},
    )
    updated = await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0})
    enriched = await _enrich_with_stagiaire([updated])
    return enriched[0]


@router.post("/library/auto-attach")
async def auto_attach_factures(user: dict = Depends(deps.get_current_user)):
    """Parcourt toutes les factures non-rattachées et tente un auto-rattachement.

    Pour chaque PDF facture orphelin :
      1. Normalise le n° depuis le nom de fichier
      2. Cherche la facture CPF correspondante
      3. Récupère le n° de dossier CPF
      4. Cherche l'apprenant avec ce dossier_cpf
      5. Rattache si tout matche

    Retourne un rapport détaillé (succès + anomalies).
    """
    # Filtre : factures non rattachées à un apprenant ET PDF/Image (pas xlsx ni rapports)
    query = {
        "type": "facture",
        "$or": [
            {"apprenant_id": None},
            {"apprenant_id": {"$exists": False}},
        ],
    }
    docs = await deps.db.dossier_documents.find(query, {"_id": 0}).to_list(10000)

    report = {
        "total_examined": len(docs),
        "attached": 0,
        "already_attached": 0,
        "skipped": [],          # fichiers ignorés (format non reconnu, extension non PDF, etc.)
        "anomalies": [],        # facture introuvable, dossier orphelin
        "successes": [],        # liste des rattachements effectués
    }

    for doc in docs:
        original = doc.get("original_filename") or doc.get("filename") or ""
        ext = Path(original).suffix.lower().lstrip(".")
        # Ignore les fichiers non-facture (Excel, rapports, etc.)
        if ext in ("xlsx", "xls", "xlsm", "csv", "doc", "docx"):
            report["skipped"].append({
                "id": doc.get("id"),
                "filename": original,
                "reason": f"Format {ext.upper()} ignoré (auto-rattachement réservé aux PDF/images de factures)",
            })
            continue

        match = await _match_apprenant_for_filename(original)
        status = match.get("status")
        if status == "ok":
            await deps.db.dossier_documents.update_one(
                {"id": doc["id"]},
                {"$set": {
                    "apprenant_id": match["apprenant_id"],
                    "auto_attached": True,
                    "auto_attach_meta": {
                        "numero_facture": match.get("numero_facture"),
                        "numero_dossier": match.get("numero_dossier"),
                    },
                }},
            )
            report["attached"] += 1
            report["successes"].append({
                "id": doc.get("id"),
                "filename": original,
                "apprenant": f"{match.get('apprenant_prenom', '')} {match.get('apprenant_nom', '')}".strip(),
                "numero_facture": match.get("numero_facture"),
                "numero_dossier": match.get("numero_dossier"),
            })
        elif status == "unparseable":
            report["skipped"].append({
                "id": doc.get("id"),
                "filename": original,
                "reason": match.get("reason"),
            })
        else:  # invoice_not_found, no_apprenant
            report["anomalies"].append({
                "id": doc.get("id"),
                "filename": original,
                "reason": match.get("reason"),
                "numero_facture": match.get("numero_facture"),
                "numero_dossier": match.get("numero_dossier"),
            })

    return report


@router.post("/library/auto-attach/export-xlsx")
async def export_auto_attach_report(payload: dict = Body(...), user: dict = Depends(deps.get_current_user)):
    """Convertit un rapport d'auto-rattachement en fichier Excel (3 feuilles).

    Le client renvoie le rapport JSON retourné par /library/auto-attach. Le serveur
    génère un xlsx stylé (Rattachements / Anomalies / Ignorés) à transmettre à l'équipe admin.
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="left", vertical="center")

    def _add_sheet(name: str, headers: list, rows: list, header_color: str):
        ws = wb.create_sheet(name)
        fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = fill
            c.alignment = header_align
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
        # Ajuste largeur colonnes (auto-ish)
        for col_idx, h in enumerate(headers, start=1):
            max_len = max([len(str(h))] + [len(str(r[col_idx - 1] or "")) for r in rows] + [10])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
        ws.freeze_panes = "A2"

    # Feuille 1 : Synthèse
    ws_syn = wb.create_sheet("Synthèse")
    ws_syn["A1"] = "Rapport d'auto-rattachement des factures"
    ws_syn["A1"].font = Font(bold=True, size=14)
    ws_syn["A3"] = "Documents analysés"
    ws_syn["B3"] = payload.get("total_examined", 0)
    ws_syn["A4"] = "Rattachements effectués"
    ws_syn["B4"] = payload.get("attached", 0)
    ws_syn["A5"] = "Anomalies (à traiter manuellement)"
    ws_syn["B5"] = len(payload.get("anomalies", []))
    ws_syn["A6"] = "Fichiers ignorés (format non reconnu)"
    ws_syn["B6"] = len(payload.get("skipped", []))
    for row in range(3, 7):
        ws_syn.cell(row=row, column=1).font = Font(bold=True)
    ws_syn.column_dimensions["A"].width = 40
    ws_syn.column_dimensions["B"].width = 15

    # Feuille 2 : Rattachements
    _add_sheet(
        "Rattachements",
        ["Fichier", "Apprenant", "N° facture", "N° dossier CPF"],
        [[s.get("filename"), s.get("apprenant"), s.get("numero_facture"), s.get("numero_dossier")] for s in payload.get("successes", [])],
        "10B981",  # emerald
    )

    # Feuille 3 : Anomalies
    _add_sheet(
        "Anomalies",
        ["Fichier", "Motif", "N° facture détecté", "N° dossier détecté"],
        [[a.get("filename"), a.get("reason"), a.get("numero_facture"), a.get("numero_dossier")] for a in payload.get("anomalies", [])],
        "DC2626",  # red
    )

    # Feuille 4 : Ignorés
    _add_sheet(
        "Ignorés",
        ["Fichier", "Motif"],
        [[s.get("filename"), s.get("reason")] for s in payload.get("skipped", [])],
        "64748B",  # slate
    )

    buf = _io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    filename = f"rapport_rattachement_factures_{deps.now_utc().strftime('%Y-%m-%d_%H%M')}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(content)),
            "Cache-Control": "no-store",
        },
    )


@router.patch("/library/{document_id}/detach")
async def detach_document(document_id: str, user: dict = Depends(deps.get_current_user)):
    if not await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0, "id": 1}):
        raise HTTPException(404, "Document introuvable")
    await deps.db.dossier_documents.update_one({"id": document_id}, {"$set": {"dossier_id": None}})
    updated = await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0})
    enriched = await _enrich_with_stagiaire([updated])
    return enriched[0]


@router.patch("/library/{document_id}")
async def update_document(document_id: str, payload: dict = Body(...), user: dict = Depends(deps.get_current_user)):
    updates = {}
    if "type" in payload and payload["type"] in DOC_TYPE_LABEL:
        updates["type"] = payload["type"]
    if not updates:
        raise HTTPException(400, "Aucune mise à jour valide")
    res = await deps.db.dossier_documents.update_one({"id": document_id}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "Document introuvable")
    updated = await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0})
    enriched = await _enrich_with_stagiaire([updated])
    return enriched[0]


@router.delete("/library/bulk")
async def delete_library_bulk(payload: dict = Body(...), user: dict = Depends(deps.get_current_user)):
    ids = payload.get("ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, "Liste 'ids' requise et non vide")
    res = await deps.db.dossier_documents.delete_many({"id": {"$in": ids}})
    return {"deleted": res.deleted_count}


@router.delete("/library/{document_id}")
async def delete_library_doc(document_id: str, user: dict = Depends(deps.get_current_user)):
    doc = await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document introuvable")
    await deps.db.dossier_documents.delete_one({"id": document_id})
    return {"deleted": 1}


@router.get("/library/{document_id}/download")
async def download_library_doc(document_id: str, user: dict = Depends(deps.get_current_user)):
    doc = await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document introuvable")
    data, ctype = await _fetch_bytes(doc)
    filename = doc.get("original_filename", doc.get("filename", "document"))
    return Response(
        content=data,
        media_type=ctype,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(data)),
            "Cache-Control": "no-store",
        },
    )


@router.get("/library/{document_id}/preview")
async def preview_library_doc(document_id: str, user: dict = Depends(deps.get_current_user)):
    doc = await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document introuvable")
    data, ctype = await _fetch_bytes(doc)
    filename = doc.get("original_filename", doc.get("filename", "document"))
    return Response(
        content=data,
        media_type=ctype,
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Content-Length": str(len(data)),
            "Cache-Control": "no-store",
        },
    )


@router.get("/library/{document_id}/preview-html")
async def preview_library_html(document_id: str, user: dict = Depends(deps.get_current_user)):
    """Aperçu HTML pour les Excel/CSV (rendu inline du contenu)."""
    import csv as _csv
    import io as _io
    from html import escape as _esc

    doc = await deps.db.dossier_documents.find_one({"id": document_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document introuvable")
    data, _ctype = await _fetch_bytes(doc)

    ext = (doc.get("original_filename") or doc.get("filename") or "").lower().split(".")[-1]
    sheets_html = []
    MAX_ROWS, MAX_COLS = 500, 40

    try:
        if ext in ("xlsx", "xlsm", "xls"):
            from openpyxl import load_workbook
            wb = load_workbook(filename=_io.BytesIO(data), data_only=True, read_only=True)
            for ws in wb.worksheets:
                rows_html = []
                first = True
                count = 0
                for row in ws.iter_rows(values_only=True):
                    if count >= MAX_ROWS:
                        rows_html.append(f'<tr><td colspan="{MAX_COLS}" class="trunc">… ({MAX_ROWS} lignes affichées sur plus)</td></tr>')
                        break
                    cells = list(row)[:MAX_COLS]
                    if first:
                        rows_html.append("<thead><tr>" + "".join(f"<th>{_esc(str(c) if c is not None else '')}</th>" for c in cells) + "</tr></thead><tbody>")
                        first = False
                    else:
                        rows_html.append("<tr>" + "".join(f"<td>{_esc(str(c) if c is not None else '')}</td>" for c in cells) + "</tr>")
                    count += 1
                if rows_html:
                    rows_html.append("</tbody>")
                else:
                    rows_html.append('<tbody><tr><td class="empty">Feuille vide</td></tr></tbody>')
                sheets_html.append({"name": ws.title, "html": '<table class="xls-table">' + "".join(rows_html) + "</table>"})
            wb.close()
        elif ext == "csv":
            try:
                text = data.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = data.decode("latin-1", errors="ignore")
            sample = text[:4096]
            delim = ";" if sample.count(";") > sample.count(",") else ","
            reader = _csv.reader(text.splitlines(), delimiter=delim)
            rows_html = []
            for i, row in enumerate(reader):
                if i >= MAX_ROWS:
                    rows_html.append(f'<tr><td colspan="{MAX_COLS}" class="trunc">… ({MAX_ROWS} lignes affichées sur plus)</td></tr>')
                    break
                cells = row[:MAX_COLS]
                if i == 0:
                    rows_html.append("<thead><tr>" + "".join(f"<th>{_esc(c)}</th>" for c in cells) + "</tr></thead><tbody>")
                else:
                    rows_html.append("<tr>" + "".join(f"<td>{_esc(c)}</td>" for c in cells) + "</tr>")
            rows_html.append("</tbody>")
            sheets_html.append({"name": "CSV", "html": '<table class="xls-table">' + "".join(rows_html) + "</table>"})
        else:
            raise HTTPException(400, f"Aperçu HTML non disponible pour ce format ({ext})")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(500, f"Erreur de lecture du fichier : {exc}")

    return {"filename": doc.get("original_filename"), "sheets": sheets_html, "total_sheets": len(sheets_html)}
