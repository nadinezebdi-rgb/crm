"""Export EDOF (CSV / XLSX), Génération PDF, Extraction PDF via LLM."""
import io
import csv
import json
import re
import os
from pathlib import Path
from typing import Optional, List
from datetime import datetime

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pypdf import PdfReader
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

import deps

router = APIRouter()

NIVEAUX_ANGLAIS = ["Anglais A1", "Anglais A2", "Anglais B1", "Anglais B2", "Anglais C1", "Anglais C2"]


# ============================================================================
# EXPORT EDOF — CSV / XLSX
# ============================================================================

EXPORT_COLUMNS = [
    ("nom", "Nom"),
    ("prenom", "Prénom"),
    ("date_naissance", "Date de naissance"),
    ("adresse", "Adresse"),
    ("email", "Email"),
    ("telephone", "Téléphone"),
    ("formation", "Formation"),
    ("financeur_type", "Type Financeur"),
    ("financeur_nom", "Détail Financeur"),
    ("formateur_nom", "Formateur"),
    ("date_entree", "Date d'entrée"),
    ("date_debut_formation", "Début formation"),
    ("date_fin_formation", "Fin formation"),
    ("status", "Statut"),
    ("date_cloture", "Date clôture"),
    ("notes", "Notes"),
]

STATUS_LABEL = {
    "devis_attente": "Devis en attente",
    "devis_valide": "Devis validé",
    "en_formation": "En action de formation",
    "fin_formation": "Fin d'action de formation",
    "facture": "Facturé",
    "regle": "Réglé (clôturé)",
}


def _fmt_date(s: Optional[str]) -> str:
    if not s:
        return ""
    try:
        d = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return d.strftime("%d/%m/%Y")
    except Exception:
        return s


def _cell(d: dict, key: str) -> str:
    v = d.get(key)
    if v is None:
        return ""
    if key in ("date_entree", "date_debut_formation", "date_fin_formation", "date_cloture", "date_naissance"):
        return _fmt_date(v)
    if key == "status":
        return STATUS_LABEL.get(v, v)
    return str(v)


async def _fetch_dossiers(scope: str) -> List[dict]:
    query: dict = {}
    if scope == "active":
        query = {"status": {"$ne": "regle"}}
    elif scope == "closed":
        query = {"status": "regle"}
    items = await deps.db.dossiers.find(query, {"_id": 0}).sort("created_at", -1).to_list(50000)
    # Attache le nom du formateur
    for d in items:
        fid = d.get("formateur_id")
        if fid:
            f = await deps.db.formateurs.find_one({"id": fid}, {"_id": 0})
            if f:
                d["formateur_nom"] = f"{f.get('prenom', '')} {f.get('nom', '')}".strip()
    return items


@router.get("/dossiers-admin/export")
async def export_dossiers(
    format: str = Query("xlsx", regex="^(csv|xlsx)$"),
    scope: str = Query("all", regex="^(all|active|closed)$"),
    user: dict = Depends(deps.get_current_user),
):
    """Télécharge un export EDOF des dossiers (CSV ou Excel)."""
    rows = await _fetch_dossiers(scope)

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow([label for _, label in EXPORT_COLUMNS])
        for d in rows:
            writer.writerow([_cell(d, key) for key, _ in EXPORT_COLUMNS])
        data = buf.getvalue().encode("utf-8-sig")  # BOM pour Excel
        filename = f"export_edof_{scope}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Dossiers"
    ws.append([label for _, label in EXPORT_COLUMNS])
    # Style entête
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for d in rows:
        ws.append([_cell(d, key) for key, _ in EXPORT_COLUMNS])
    # Largeur colonnes
    widths = [12, 12, 14, 30, 26, 14, 22, 12, 22, 22, 14, 14, 14, 22, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"export_edof_{scope}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================================
# PDF — Génération d'une fiche dossier
# ============================================================================


@router.get("/dossiers/{dossier_id}/pdf")
async def dossier_pdf(dossier_id: str, user: dict = Depends(deps.get_current_user)):
    """Génère et télécharge un PDF récapitulatif du dossier."""
    d = await deps.db.dossiers.find_one({"id": dossier_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dossier introuvable")

    # Attache formateur
    fid = d.get("formateur_id")
    if fid:
        f = await deps.db.formateurs.find_one({"id": fid}, {"_id": 0})
        if f:
            d["formateur_nom"] = f"{f.get('prenom', '')} {f.get('nom', '')}".strip()

    # Documents associés
    documents = await deps.db.dossier_documents.find(
        {"dossier_id": dossier_id}, {"_id": 0}
    ).sort("uploaded_at", -1).to_list(200)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.8 * cm, rightMargin=1.8 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=18, leading=22, textColor=colors.HexColor("#0F172A"), spaceAfter=4)
    subtitle_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"), spaceAfter=18, leading=12)
    section_style = ParagraphStyle("section", parent=styles["Heading2"], fontSize=11, textColor=colors.HexColor("#0F172A"), spaceBefore=12, spaceAfter=8, leading=14)

    story = []
    story.append(Paragraph(f"Dossier stagiaire — {d.get('prenom', '')} {d.get('nom', '')}", title_style))
    story.append(Paragraph(
        f"Statut : <b>{STATUS_LABEL.get(d.get('status'), d.get('status') or '—')}</b> &nbsp;·&nbsp; "
        f"Financeur : <b>{d.get('financeur_type') or '—'}</b> &nbsp;·&nbsp; "
        f"Dossier #{(d.get('id') or '')[:8]}",
        subtitle_style,
    ))

    def info_table(title: str, pairs: list) -> Table:
        data = [[Paragraph(f"<b>{k}</b>", styles["Normal"]), Paragraph(str(v) if v else "—", styles["Normal"])] for k, v in pairs]
        tbl = Table(data, colWidths=[5 * cm, 11 * cm])
        tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        return tbl

    story.append(Paragraph("Identité", section_style))
    story.append(info_table("Identité", [
        ("Nom", d.get("nom")),
        ("Prénom", d.get("prenom")),
        ("Date de naissance", _fmt_date(d.get("date_naissance"))),
        ("Adresse", d.get("adresse")),
        ("Email", d.get("email")),
        ("Téléphone", d.get("telephone")),
    ]))

    story.append(Paragraph("Formation", section_style))
    story.append(info_table("Formation", [
        ("Intitulé", d.get("formation")),
        ("Formateur attribué", d.get("formateur_nom")),
        ("Date d'entrée", _fmt_date(d.get("date_entree"))),
        ("Début formation", _fmt_date(d.get("date_debut_formation"))),
        ("Fin formation", _fmt_date(d.get("date_fin_formation"))),
        ("Date de clôture", _fmt_date(d.get("date_cloture"))),
    ]))

    story.append(Paragraph("Financement", section_style))
    story.append(info_table("Financement", [
        ("Type de financeur", d.get("financeur_type")),
        ("Détail financeur", d.get("financeur_nom")),
    ]))

    if d.get("notes"):
        story.append(Paragraph("Notes", section_style))
        story.append(Paragraph(d.get("notes"), styles["Normal"]))

    if documents:
        story.append(Paragraph(f"Documents archivés ({len(documents)})", section_style))
        rows_doc = [["Type", "Fichier", "Importé le"]]
        type_label = {
            "devis_signe": "Devis signé",
            "attestation": "Attestation",
            "facture": "Facture",
            "justificatif_paiement": "Justificatif de paiement",
        }
        for doc_item in documents:
            rows_doc.append([
                type_label.get(doc_item.get("type"), doc_item.get("type", "")),
                doc_item.get("original_filename", ""),
                _fmt_date(doc_item.get("uploaded_at")),
            ])
        tbl = Table(rows_doc, colWidths=[5 * cm, 8 * cm, 3 * cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(tbl)

    story.append(Spacer(1, 30))
    story.append(Paragraph(
        f"<font color='#94A3B8' size='8'>Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — Blade Academy CRM</font>",
        styles["Normal"],
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"dossier_{d.get('nom', 'stagiaire')}_{d.get('prenom', '')}.pdf".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================================
# EXTRACTION PDF — Lit un PDF, extrait les coordonnées via LLM
# ============================================================================

_NIVEAU_RE = re.compile(r"(?:anglais\s*)?(a1|a2|b1|b2|c1|c2)\b", re.IGNORECASE)


def _detect_niveau(text: str) -> Optional[str]:
    m = _NIVEAU_RE.search(text or "")
    if m:
        return f"Anglais {m.group(1).upper()}"
    return None


def _detect_email(text: str) -> Optional[str]:
    m = re.search(r"[\w\.\-_+]+@[\w\.\-]+\.[a-zA-Z]{2,}", text or "")
    return m.group(0) if m else None


def _detect_phone(text: str) -> Optional[str]:
    # FR : 0X XX XX XX XX, +33...
    m = re.search(r"(?:\+33\s?|0)[1-9](?:[\s\.\-]?\d{2}){4}", text or "")
    return m.group(0).strip() if m else None


def _detect_date_naissance(text: str) -> Optional[str]:
    # Cherche après "né(e) le", "date de naissance", etc.
    patterns = [
        r"(?:n[ée]\(?e?\)?\s*le|date\s+de\s+naissance)\s*:?\s*(\d{1,2}[/\.\-]\d{1,2}[/\.\-]\d{2,4})",
        r"(?:n[ée]\(?e?\)?\s*le|date\s+de\s+naissance)\s*:?\s*(\d{1,2}\s+\w+\s+\d{4})",
    ]
    for p in patterns:
        m = re.search(p, text or "", re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None


async def _llm_extract(text: str) -> dict:
    """Utilise Claude Haiku pour extraire les champs structurés depuis le texte PDF."""
    import logging
    logger = logging.getLogger("crm.llm_extract")
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
    except Exception as exc:
        logger.warning(f"emergentintegrations import failed: {exc}")
        return {"_llm_error": f"import: {exc}"}

    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        logger.warning("EMERGENT_LLM_KEY not set in environment")
        return {"_llm_error": "EMERGENT_LLM_KEY not set"}

    # Limite la taille du texte transmis (LLM token budget)
    truncated = (text or "")[:8000]

    system_msg = (
        "Tu es un assistant qui extrait des informations structurées d'un dossier stagiaire PDF en français. "
        "Tu réponds UNIQUEMENT par un objet JSON valide, sans aucun texte autour, avec ces clés (toutes optionnelles) : "
        "{\"nom\": str, \"prenom\": str, \"date_naissance\": \"YYYY-MM-DD\", \"adresse\": str, "
        "\"email\": str, \"telephone\": str, \"niveau_anglais\": l'une parmi [\"Anglais A1\",\"Anglais A2\",\"Anglais B1\",\"Anglais B2\",\"Anglais C1\",\"Anglais C2\"]}. "
        "Si un champ n'est pas trouvé, mets-le à null. Pour 'niveau_anglais', repère un niveau CECRL (A1/A2/B1/B2/C1/C2). "
        "Pas de markdown, pas de bloc ```json, juste l'objet JSON brut."
    )

    try:
        chat = LlmChat(
            api_key=api_key,
            session_id=f"extract-pdf-{datetime.utcnow().timestamp()}",
            system_message=system_msg,
        ).with_model("anthropic", "claude-haiku-4-5-20251001")
        result = await chat.send_message(UserMessage(text=truncated))
        # result est une string
        if isinstance(result, str):
            # Nettoie d'éventuelles entourages markdown
            cleaned = result.strip()
            cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
            cleaned = re.sub(r"\s*```$", "", cleaned)
            data = json.loads(cleaned)
            return data or {}
        return {"_llm_error": f"unexpected result type: {type(result).__name__}"}
    except Exception as exc:
        logger.exception("LLM extract failed")
        return {"_llm_error": str(exc)}


def _parse_pdf_text(content: bytes) -> str:
    """Extrait tout le texte d'un PDF en bytes."""
    try:
        reader = PdfReader(io.BytesIO(content))
        out = []
        for page in reader.pages:
            try:
                out.append(page.extract_text() or "")
            except Exception:
                continue
        return "\n".join(out)
    except Exception as exc:
        raise HTTPException(400, f"PDF illisible : {exc}")


def _normalize_date(s: Optional[str]) -> Optional[str]:
    """Convertit en YYYY-MM-DD si possible."""
    if not s:
        return None
    s = s.strip()
    # ISO direct
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return s
    # DD/MM/YYYY ou DD-MM-YYYY
    m = re.match(r"^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        y = ("20" + y) if len(y) == 2 else y
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return s


@router.post("/dossiers/extract-pdf")
async def extract_pdf(
    file: UploadFile = File(...),
    user: dict = Depends(deps.get_current_user),
):
    """Upload un PDF, extrait les coordonnées stagiaire + niveau Anglais.
    Renvoie les champs détectés (n'enregistre rien)."""
    content = await file.read()
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(400, "Fichier PDF trop volumineux (max 15 Mo)")
    text = _parse_pdf_text(content)

    # 1) Tentative LLM
    llm_data = await _llm_extract(text) if text.strip() else {}
    # 2) Fallback regex pour combler les trous
    fallback = {
        "email": _detect_email(text),
        "telephone": _detect_phone(text),
        "date_naissance": _normalize_date(_detect_date_naissance(text)),
        "niveau_anglais": _detect_niveau(text),
    }

    merged = {}
    for key in ("nom", "prenom", "date_naissance", "adresse", "email", "telephone", "niveau_anglais"):
        v_llm = llm_data.get(key)
        v_fallback = fallback.get(key)
        # priorité au LLM s'il a trouvé qqch
        merged[key] = v_llm or v_fallback

    # Normalise la date
    merged["date_naissance"] = _normalize_date(merged.get("date_naissance"))

    # Valide niveau_anglais
    if merged.get("niveau_anglais") and merged["niveau_anglais"] not in NIVEAUX_ANGLAIS:
        merged["niveau_anglais"] = None

    # Si LLM a échoué, on l'indique sans casser
    response = {
        "extracted": merged,
        "filename": file.filename,
        "text_length": len(text),
        "llm_used": bool(llm_data) and "_llm_error" not in llm_data,
        "llm_error": llm_data.get("_llm_error"),
        "niveaux_disponibles": NIVEAUX_ANGLAIS,
    }
    return response


@router.post("/dossiers/{dossier_id}/extract-and-fill")
async def extract_and_fill(
    dossier_id: str,
    file: UploadFile = File(...),
    user: dict = Depends(deps.get_current_user),
):
    """Upload un PDF pour un dossier existant : extrait + remplit automatiquement
    les coordonnées du stagiaire et sélectionne le niveau Anglais (formation)."""
    d = await deps.db.dossiers.find_one({"id": dossier_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dossier introuvable")

    content = await file.read()
    if len(content) > 15 * 1024 * 1024:
        raise HTTPException(400, "PDF trop volumineux (max 15 Mo)")
    text = _parse_pdf_text(content)

    llm_data = await _llm_extract(text) if text.strip() else {}
    fallback = {
        "email": _detect_email(text),
        "telephone": _detect_phone(text),
        "date_naissance": _normalize_date(_detect_date_naissance(text)),
        "niveau_anglais": _detect_niveau(text),
    }

    extracted = {}
    for key in ("nom", "prenom", "date_naissance", "adresse", "email", "telephone", "niveau_anglais"):
        v = llm_data.get(key) or fallback.get(key)
        if v:
            extracted[key] = v

    extracted["date_naissance"] = _normalize_date(extracted.get("date_naissance"))
    if extracted.get("niveau_anglais") and extracted["niveau_anglais"] not in NIVEAUX_ANGLAIS:
        extracted.pop("niveau_anglais", None)

    # On ne remplit que les champs vides du dossier (préserve la donnée existante)
    updates = {}
    for key in ("nom", "prenom", "date_naissance", "adresse", "email", "telephone"):
        if extracted.get(key) and not d.get(key):
            updates[key] = extracted[key]
    # Formation : on écrase si on a un niveau plus précis
    if extracted.get("niveau_anglais"):
        updates["formation"] = extracted["niveau_anglais"]

    if updates:
        updates["updated_at"] = deps.now_utc().isoformat()
        await deps.db.dossiers.update_one({"id": dossier_id}, {"$set": updates})

    refreshed = await deps.db.dossiers.find_one({"id": dossier_id}, {"_id": 0})
    fid = refreshed.get("formateur_id") if refreshed else None
    if fid:
        f = await deps.db.formateurs.find_one({"id": fid}, {"_id": 0})
        if f:
            refreshed["formateur_nom"] = f"{f.get('prenom', '')} {f.get('nom', '')}".strip()

    return {
        "extracted": extracted,
        "applied_updates": updates,
        "dossier": refreshed,
        "llm_used": bool(llm_data) and "_llm_error" not in llm_data,
    }
