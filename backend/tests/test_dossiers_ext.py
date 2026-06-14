"""Tests for iteration 4 — Export EDOF (CSV/XLSX), PDF generation, PDF extraction."""
import os
import io
import pytest
import requests
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
from pypdf import PdfReader

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://crm-trainees.preview.emergentagent.com").rstrip("/")


@pytest.fixture(scope="module")
def auth_session():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login",
               json={"email": "admin@blade-academy.fr", "password": "admin123"},
               timeout=20)
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    return s


@pytest.fixture(scope="module")
def sample_dossier_id(auth_session):
    """Create a dossier we'll use across PDF tests, return its id; cleanup at the end."""
    formateurs = auth_session.get(f"{BASE_URL}/api/formateurs", timeout=10).json()
    fid = formateurs[0]["id"] if formateurs else None
    payload = {
        "nom": "TEST_PDF",
        "prenom": "Stagiaire",
        "email": "test_pdf@example.com",
        "telephone": "06 11 22 33 44",
        "formation": "Anglais B1",
        "financeur_type": "CPF",
        "financeur_nom": "CPF",
        "formateur_id": fid,
        "status": "devis_attente",
    }
    r = auth_session.post(f"{BASE_URL}/api/dossiers", json=payload, timeout=15)
    assert r.status_code in (200, 201), f"create failed: {r.status_code} {r.text}"
    did = r.json()["id"]
    yield did
    auth_session.delete(f"{BASE_URL}/api/dossiers/{did}", timeout=10)


def _build_test_pdf() -> bytes:
    """Generate an in-memory PDF with French stagiaire info."""
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    c.setFont("Helvetica", 12)
    y = 800
    lines = [
        "DOSSIER STAGIAIRE",
        "Nom: MARTIN",
        "Prénom: Sophie",
        "Né(e) le: 15/06/1985",
        "Adresse: 12 rue de Paris, 75001 Paris",
        "Email: sophie.martin@example.com",
        "Téléphone: 06 12 34 56 78",
        "Formation: Anglais B2",
    ]
    for line in lines:
        c.drawString(50, y, line)
        y -= 20
    c.save()
    return buf.getvalue()


# ============================================================================
# Auth gating
# ============================================================================

class TestAuthGuard:
    def test_export_requires_auth(self):
        r = requests.get(f"{BASE_URL}/api/dossiers-admin/export?format=csv&scope=all", timeout=10)
        assert r.status_code == 401

    def test_pdf_requires_auth(self):
        r = requests.get(f"{BASE_URL}/api/dossiers/anyid/pdf", timeout=10)
        assert r.status_code == 401

    def test_extract_requires_auth(self):
        r = requests.post(f"{BASE_URL}/api/dossiers/extract-pdf",
                          files={"file": ("a.pdf", b"x", "application/pdf")}, timeout=10)
        assert r.status_code == 401

    def test_extract_and_fill_requires_auth(self):
        r = requests.post(f"{BASE_URL}/api/dossiers/anyid/extract-and-fill",
                          files={"file": ("a.pdf", b"x", "application/pdf")}, timeout=10)
        assert r.status_code == 401


# ============================================================================
# EXPORT EDOF (CSV / XLSX) + scope/format validation
# ============================================================================

class TestExportEdof:
    def test_export_xlsx_all(self, auth_session):
        r = auth_session.get(f"{BASE_URL}/api/dossiers-admin/export?format=xlsx&scope=all", timeout=30)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "xlsx" in ct
        assert "attachment" in r.headers.get("content-disposition", "").lower()
        assert ".xlsx" in r.headers.get("content-disposition", "")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # 16 expected headers
        headers = [c.value for c in ws[1]]
        expected = ["Nom", "Prénom", "Date de naissance", "Adresse", "Email", "Téléphone",
                    "Formation", "Type Financeur", "Détail Financeur", "Formateur",
                    "Date d'entrée", "Début formation", "Fin formation", "Statut",
                    "Date clôture", "Notes"]
        assert headers == expected
        # Header styling: white text on #0F172A
        c0 = ws[1][0]
        assert c0.font.bold is True
        assert (c0.font.color.rgb or "").upper().endswith("FFFFFF")
        assert (c0.fill.start_color.rgb or "").upper().endswith("0F172A")

    def test_export_csv_has_bom_and_semicolon(self, auth_session):
        r = auth_session.get(f"{BASE_URL}/api/dossiers-admin/export?format=csv&scope=all", timeout=30)
        assert r.status_code == 200
        assert "text/csv" in r.headers.get("content-type", "")
        # BOM UTF-8
        assert r.content[:3] == b"\xef\xbb\xbf"
        text = r.content.decode("utf-8-sig")
        first_line = text.splitlines()[0]
        assert ";" in first_line
        assert "Nom" in first_line and "Prénom" in first_line

    def test_export_scope_active_excludes_regle(self, auth_session):
        r_active = auth_session.get(f"{BASE_URL}/api/dossiers-admin/export?format=csv&scope=active", timeout=30)
        r_closed = auth_session.get(f"{BASE_URL}/api/dossiers-admin/export?format=csv&scope=closed", timeout=30)
        assert r_active.status_code == 200 and r_closed.status_code == 200
        a_text = r_active.content.decode("utf-8-sig").splitlines()
        c_text = r_closed.content.decode("utf-8-sig").splitlines()
        # No row of active should have "Réglé (clôturé)"
        for row in a_text[1:]:
            assert "Réglé (clôturé)" not in row
        # All rows of closed should have "Réglé (clôturé)"
        for row in c_text[1:]:
            assert "Réglé (clôturé)" in row

    def test_invalid_format_422(self, auth_session):
        r = auth_session.get(f"{BASE_URL}/api/dossiers-admin/export?format=pdf&scope=all", timeout=10)
        assert r.status_code == 422

    def test_invalid_scope_422(self, auth_session):
        r = auth_session.get(f"{BASE_URL}/api/dossiers-admin/export?format=csv&scope=archived", timeout=10)
        assert r.status_code == 422


# ============================================================================
# PDF generation
# ============================================================================

class TestDossierPdf:
    def test_pdf_404_unknown(self, auth_session):
        r = auth_session.get(f"{BASE_URL}/api/dossiers/does-not-exist/pdf", timeout=10)
        assert r.status_code == 404

    def test_pdf_valid_for_existing(self, auth_session, sample_dossier_id):
        r = auth_session.get(f"{BASE_URL}/api/dossiers/{sample_dossier_id}/pdf", timeout=30)
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith("application/pdf")
        assert r.content[:4] == b"%PDF"
        assert len(r.content) > 1000
        # parse
        reader = PdfReader(io.BytesIO(r.content))
        text = "".join((p.extract_text() or "") for p in reader.pages)
        assert "TEST_PDF" in text or "Stagiaire" in text
        assert "Identité" in text
        assert "Formation" in text
        assert "Financement" in text


# ============================================================================
# PDF extraction (Claude Haiku + regex fallback)
# ============================================================================

class TestExtractPdf:
    def test_extract_pdf_success(self, auth_session):
        pdf_bytes = _build_test_pdf()
        r = auth_session.post(
            f"{BASE_URL}/api/dossiers/extract-pdf",
            files={"file": ("stagiaire.pdf", pdf_bytes, "application/pdf")},
            timeout=60,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert "extracted" in body
        assert "llm_used" in body
        assert "filename" in body
        assert "text_length" in body
        assert body["niveaux_disponibles"] == [
            "Anglais A1", "Anglais A2", "Anglais B1", "Anglais B2", "Anglais C1", "Anglais C2"
        ]
        ex = body["extracted"]
        # Regex fallbacks must catch at minimum email/phone/niveau even if LLM is offline
        assert ex.get("email") == "sophie.martin@example.com"
        assert "06" in (ex.get("telephone") or "")
        assert ex.get("niveau_anglais") == "Anglais B2"
        # date_naissance normalized to YYYY-MM-DD
        if ex.get("date_naissance"):
            assert ex["date_naissance"].startswith("1985-06-15") or ex["date_naissance"] == "1985-06-15"

    def test_extract_pdf_too_big(self, auth_session):
        # 16 MB junk -> 400
        big = b"%PDF-1.4\n" + b"0" * (16 * 1024 * 1024)
        r = auth_session.post(
            f"{BASE_URL}/api/dossiers/extract-pdf",
            files={"file": ("big.pdf", big, "application/pdf")},
            timeout=60,
        )
        assert r.status_code == 400

    def test_extract_pdf_invalid_format(self, auth_session):
        r = auth_session.post(
            f"{BASE_URL}/api/dossiers/extract-pdf",
            files={"file": ("notpdf.txt", b"this is not a pdf", "text/plain")},
            timeout=15,
        )
        # _parse_pdf_text raises 400 'PDF illisible'
        assert r.status_code == 400

    def test_extract_and_fill_404(self, auth_session):
        pdf_bytes = _build_test_pdf()
        r = auth_session.post(
            f"{BASE_URL}/api/dossiers/no-such-dossier/extract-and-fill",
            files={"file": ("x.pdf", pdf_bytes, "application/pdf")},
            timeout=30,
        )
        assert r.status_code == 404

    def test_extract_and_fill_applies_only_empty_fields(self, auth_session):
        # Create a dossier with email already filled, formation = old value
        formateurs = auth_session.get(f"{BASE_URL}/api/formateurs", timeout=10).json()
        fid = formateurs[0]["id"] if formateurs else None
        create = auth_session.post(f"{BASE_URL}/api/dossiers", json={
            "nom": "TEST_FILL",
            "prenom": "Bob",
            "email": "preexisting@example.com",
            "formation": "Anglais A1",
            "financeur_type": "CPF",
            "financeur_nom": "CPF",
            "formateur_id": fid,
            "status": "devis_attente",
        }, timeout=15)
        assert create.status_code in (200, 201), create.text
        did = create.json()["id"]
        try:
            pdf_bytes = _build_test_pdf()  # contains email sophie.martin / niveau B2
            r = auth_session.post(
                f"{BASE_URL}/api/dossiers/{did}/extract-and-fill",
                files={"file": ("s.pdf", pdf_bytes, "application/pdf")},
                timeout=60,
            )
            assert r.status_code == 200, r.text
            body = r.json()
            assert "extracted" in body
            assert "applied_updates" in body
            assert "dossier" in body
            applied = body["applied_updates"]
            # Email was already set -> must NOT have been overwritten
            assert "email" not in applied
            assert body["dossier"]["email"] == "preexisting@example.com"
            # Formation IS overwritten because niveau Anglais was detected
            assert applied.get("formation") == "Anglais B2"
            assert body["dossier"]["formation"] == "Anglais B2"
            # Telephone was empty -> should be filled
            if applied.get("telephone"):
                assert "06" in applied["telephone"]
        finally:
            auth_session.delete(f"{BASE_URL}/api/dossiers/{did}", timeout=10)


# ============================================================================
# Regression — pre-existing endpoints still working after litellm/emergentintegrations cleanup
# ============================================================================

class TestRegression:
    def test_active_dossiers(self, auth_session):
        r = auth_session.get(f"{BASE_URL}/api/dossiers/active", timeout=10)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_sessions(self, auth_session):
        r = auth_session.get(f"{BASE_URL}/api/sessions", timeout=10)
        assert r.status_code == 200

    def test_formateurs(self, auth_session):
        r = auth_session.get(f"{BASE_URL}/api/formateurs", timeout=10)
        assert r.status_code == 200
        names = {f.get("prenom", "").upper() for f in r.json()}
        # 3 seeded formateurs
        assert {"NEO", "HIGH", "VIRGINIA"}.issubset(names)

    def test_clear_endpoint_still_exists(self, auth_session):
        # We don't actually clear; just verify endpoint returns 200 with active scope (likely 0 / does nothing destructive when scope=closed is empty)
        # safer: send scope=closed which currently has no rows in test env
        r = auth_session.delete(f"{BASE_URL}/api/dossiers-admin/clear?scope=closed", timeout=15)
        assert r.status_code in (200, 204)

    def test_import_edof_endpoint_exists(self, auth_session):
        # No file -> 422
        r = auth_session.post(f"{BASE_URL}/api/dossiers-admin/import-edof", timeout=10)
        assert r.status_code in (400, 422)
