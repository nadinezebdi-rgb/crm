"""Blade Academy CRM — Import des stagiaires depuis un export EDOF / Mon Compte Formation (CPF).

Parsing tolérant : CSV (séparateur ; , ou tabulation, encodages utf-8/cp1252)
et Excel (.xlsx). Détection automatique des colonnes EDOF standard, avec
mappage corrigeable côté interface.
"""

import csv
import io
import re
import unicodedata
from datetime import datetime

TARGET_FIELDS = [
    {"key": "nom", "label": "Nom", "required": True},
    {"key": "prenom", "label": "Prénom", "required": True},
    {"key": "email", "label": "Email", "required": False},
    {"key": "telephone", "label": "Téléphone", "required": False},
    {"key": "dossier", "label": "N° de dossier CPF", "required": False},
    {"key": "formation", "label": "Intitulé de la formation", "required": False},
    {"key": "date_debut", "label": "Date de début de session", "required": False},
    {"key": "date_fin", "label": "Date de fin de session", "required": False},
    {"key": "prix", "label": "Prix / montant", "required": False},
    {"key": "statut", "label": "Statut du dossier", "required": False},
]


def _norm(s):
    s = unicodedata.normalize("NFD", str(s or "").lower().strip())
    return "".join(ch for ch in s if unicodedata.category(ch) != "Mn")


# Mots-clés par champ, du plus spécifique au plus générique.
_KEYWORDS = {
    "prenom": ["prenom"],
    "nom": ["nom de naissance", "nom d'usage", "nom dusage", "nom du titulaire", "nom"],
    "email": ["mail", "courriel", "adresse electronique"],
    "telephone": ["telephone", "portable", "mobile", "tel"],
    "dossier": ["dossier"],
    "formation": ["intitule de la formation", "intitule de l'action", "intitule", "libelle de la formation", "action de formation", "formation"],
    "date_debut": ["date de debut", "debut de session", "date d'entree", "entree en formation", "date debut"],
    "date_fin": ["date de fin", "fin de session", "date de sortie", "sortie de formation", "date fin"],
    "prix": ["prix de vente", "prix", "montant", "cout"],
    "statut": ["statut", "etat"],
}

# Colonnes interdites pour certains champs (évite "prénom" capté par "nom",
# ou "date d'entrée en formation" capté par "formation").
_EXCLUDE = {
    "nom": ["prenom", "organisme", "formation", "session"],
    "formation": ["date", "entree", "sortie", "debut", "fin", "numero", "n°"],
    "telephone": ["intitule"],
    "prix": ["heure"],
}


def auto_map(columns):
    used = set()
    mapping = {}
    for field in ["prenom", "nom", "email", "telephone", "dossier", "formation", "date_debut", "date_fin", "prix", "statut"]:
        mapping[field] = None
        for kw in _KEYWORDS[field]:
            found = None
            for col in columns:
                if col in used:
                    continue
                cn = _norm(col)
                if kw in cn and not any(x in cn for x in _EXCLUDE.get(field, [])):
                    found = col
                    break
            if found:
                mapping[field] = found
                used.add(found)
                break
    return mapping


def _stringify(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def parse_import_file(filename: str, content: bytes):
    """Retourne (columns, rows) — rows = liste de dicts {colonne: str}."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_xlsx(content: bytes):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise ValueError("Fichier Excel illisible. Exportez au format .xlsx ou .csv.")
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    headers = None
    rows = []
    for raw in rows_iter:
        values = [_stringify(v) for v in raw]
        if headers is None:
            if any(values):
                headers = [v or f"Colonne {i + 1}" for i, v in enumerate(values)]
            continue
        if not any(values):
            continue
        rows.append({headers[i]: (values[i] if i < len(values) else "") for i in range(len(headers))})
    wb.close()
    if not headers:
        raise ValueError("Aucun en-tête détecté dans le fichier Excel.")
    return headers, rows


def _parse_csv(content: bytes):
    text = None
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Encodage du fichier non reconnu.")
    first_line = text.splitlines()[0] if text.splitlines() else ""
    delimiter = max([";", ",", "\t"], key=first_line.count)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise ValueError("Aucun en-tête détecté dans le fichier CSV.")
    headers = [h.strip() for h in reader.fieldnames if h and h.strip()]
    rows = []
    for raw in reader:
        row = {(k or "").strip(): _stringify(v) for k, v in raw.items() if k and k.strip()}
        if any(row.values()):
            rows.append(row)
    return headers, rows


def parse_date_fr(value):
    """'25/03/2026', '2026-03-25', '25/03/26' → ISO 'YYYY-MM-DD' ou None."""
    v = str(value or "").strip()
    if not v:
        return None
    v = v.split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(v, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_amount(value):
    """'1 495,00 €' / '1495.00' → float."""
    v = re.sub(r"[^\d,.\-]", "", str(value or ""))
    if not v:
        return 0.0
    if "," in v and "." in v:
        v = v.replace(".", "").replace(",", ".")
    elif "," in v:
        v = v.replace(",", ".")
    try:
        return float(v)
    except ValueError:
        return 0.0


# ---------------------------------------------------------------------------
# Export EDOF "Factures" (encaissements CPF)
# ---------------------------------------------------------------------------
_FACTURE_KEYWORDS = {
    "numero_facture": (["numero_facture", "numero facture", "n° facture"], ["type", "date", "emission"]),
    "numero_dossier": (["numero_dossier", "numero dossier", "dossier"], ["controle"]),
    "type_facture": (["type"], []),
    "date_emission": (["emission"], []),
    "montant": (["montant", "prix"], []),
    "statut_reglement": (["statut"], []),
    "date_reglement": (["date_reglement", "date reglement", "reglement"], ["statut", "montant"]),
    "en_controle": (["controle"], []),
}


def map_facture_columns(columns):
    used = set()
    mapping = {}
    for field, (keywords, excludes) in _FACTURE_KEYWORDS.items():
        mapping[field] = None
        for kw in keywords:
            found = None
            for col in columns:
                if col in used:
                    continue
                cn = _norm(col)
                if kw in cn and not any(x in cn for x in excludes):
                    found = col
                    break
            if found:
                mapping[field] = found
                used.add(found)
                break
    return mapping


NIVEAUX_CECRL = ["A1", "A2", "B1", "B2", "C1", "C2"]

def detect_niveau_anglais(formation):
    """Renvoie le niveau CECRL (A1..C2) si l'intitulé contient « anglais », sinon None."""
    f = str(formation or "")
    if "anglais" not in f.lower():
        return None
    m = re.search(r"\b([abcABC][12])\b", f)
    return m.group(1).upper() if m else None
